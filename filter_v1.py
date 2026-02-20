from openpyxl import load_workbook, Workbook
from pathlib import Path
import argparse


def parse_args():
    p = argparse.ArgumentParser(
        prog="filter_version",
        description="Filtra un Excel de endpoints por Version (v1/v2) y genera un archivo nuevo."
    )
    p.add_argument("--src", required=True, help="Ruta del Excel origen (EndPoints_DISCOVERED.xlsx)")
    p.add_argument("--dst", required=True, help="Ruta del Excel destino (EndPoints_DISCOVERED_v1.xlsx / v2.xlsx)")
    p.add_argument("--version", required=True, choices=["v1", "v2"], help="Version a filtrar: v1 o v2")
    p.add_argument("--sheet", default=None, help="Nombre de la hoja (opcional). Si no se indica, usa la primera.")
    p.add_argument("--column", default="Version", help="Nombre de la columna de version (default: Version)")
    return p.parse_args()


def main():
    args = parse_args()

    src = Path(args.src).expanduser().resolve()
    dst = Path(args.dst).expanduser().resolve()
    version = args.version.lower().strip()
    col_name = args.column.strip()

    if not src.exists():
        raise FileNotFoundError(f"No existe el archivo origen: {src}")

    wb = load_workbook(src)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    # Header
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    hdr = [c.value for c in header_row]

    if col_name not in hdr:
        raise ValueError(f"No se encontró la columna '{col_name}' en el Excel. Columnas: {hdr}")

    ix = hdr.index(col_name)  # 0-based index

    out = Workbook()
    ows = out.active
    ows.title = ws.title
    ows.append(hdr)

    kept = 0
    scanned = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue

        scanned += 1
        v = row[ix]
        if v is None:
            continue

        if str(v).strip().lower() == version:
            ows.append(list(row))
            kept += 1

    # Crear carpeta destino si no existe
    dst.parent.mkdir(parents=True, exist_ok=True)
    out.save(dst)

    print(f"OK -> {dst}")
    print(f"Sheet: {ws.title}")
    print(f"Scanned rows: {scanned}")
    print(f"Kept rows ({version}): {kept}")


if __name__ == "__main__":
    main()