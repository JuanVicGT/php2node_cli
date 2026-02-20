from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any

from openpyxl import load_workbook

from .utils import norm_endpoint_path, norm_http, norm_version


@dataclass
class InventoryRow:
    version: str
    controller: str
    http_method: str
    method: str
    method_base: str
    endpoint_path: str
    endpoint_path_with_version: Optional[str]
    raw: Dict[str, Any]


class Inventory:
    def __init__(self, rows: List[InventoryRow]):
        self.rows = rows

    @staticmethod
    def load(xlsx_path: Path, sheet_name: str = "EndPoints") -> "Inventory":
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]

        headers: Dict[str, int] = {}
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for idx, val in enumerate(header_row, start=1):
            if val is None:
                continue
            key = str(val).strip()
            if key:
                headers[key] = idx

        def get(row_vals, col_name: str):
            idx = headers.get(col_name)
            if not idx:
                return None
            v = row_vals[idx - 1]
            if isinstance(v, str):
                v = v.strip()
            return v

        # Formato A (completo): Version, Controller, HttpMethod, Method, MethodBase, EndpointPath...
        required_full = ["Version", "Controller", "HttpMethod", "Method", "MethodBase", "EndpointPath"]
        has_full = all(c in headers for c in required_full)

        # Formato B (mínimo): "Método"/"Metodo" + "Path"
        required_min_1 = ["Método", "Path"]
        required_min_2 = ["Metodo", "Path"]
        has_min = all(c in headers for c in required_min_1) or all(c in headers for c in required_min_2)

        if not has_full and not has_min:
            raise ValueError(
                f"Unsupported inventory format. Headers found: {list(headers.keys())}. "
                f"Expected either {required_full} or {required_min_1}."
            )

        rows: List[InventoryRow] = []

        # ---- Formato A (completo) ----
        if has_full:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                version = get(row, "Version")
                controller = get(row, "Controller")
                http_method = get(row, "HttpMethod")
                method = get(row, "Method")
                method_base = get(row, "MethodBase")
                endpoint_path = get(row, "EndpointPath")
                ep_with_ver = get(row, "EndpointPath with version")

                if not (version and controller and http_method and method and method_base and endpoint_path):
                    continue

                v = norm_version(str(version))
                hm = norm_http(str(http_method))
                ep = norm_endpoint_path(str(endpoint_path))
                epv = norm_endpoint_path(str(ep_with_ver)) if ep_with_ver else None

                rows.append(
                    InventoryRow(
                        version=v,
                        controller=str(controller).strip(),
                        http_method=hm,
                        method=str(method).strip(),
                        method_base=str(method_base).strip(),
                        endpoint_path=ep,
                        endpoint_path_with_version=epv,
                        raw={
                            "Version": v,
                            "Controller": str(controller).strip(),
                            "HttpMethod": hm,
                            "Method": str(method).strip(),
                            "MethodBase": str(method_base).strip(),
                            "EndpointPath": ep,
                            "EndpointPath with version": epv,
                            "SourceFormat": "FULL",
                        },
                    )
                )

            return Inventory(rows)

        # ---- Formato B (mínimo: Método + Path) ----
        metodo_col = "Método" if "Método" in headers else "Metodo"

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            hm = get(row, metodo_col)
            path = get(row, "Path")
            if not hm or not path:
                continue

            hm = norm_http(str(hm))
            full_path = norm_endpoint_path(str(path))

            # Infer version if path starts with v1/ or v2/
            parts = full_path.split("/")
            if parts and parts[0].lower() in ("v1", "v2"):
                inferred_version = norm_version(parts[0])
                rest = "/".join(parts[1:])
            else:
                inferred_version = "v1"
                rest = full_path

            rest = norm_endpoint_path(rest)
            rest_parts = rest.split("/")
            if len(rest_parts) < 2:
                continue

            controller = rest_parts[0]
            method_base = rest_parts[1]

            suffix = {"GET": "get", "POST": "post", "PUT": "put", "DELETE": "delete"}[hm]
            method = f"{method_base}_{suffix}"

            rows.append(
                InventoryRow(
                    version=inferred_version,
                    controller=controller,
                    http_method=hm,
                    method=method,
                    method_base=method_base,
                    endpoint_path=rest,
                    endpoint_path_with_version=full_path,
                    raw={
                        "Version": inferred_version,
                        "Controller": controller,
                        "HttpMethod": hm,
                        "Method": method,
                        "MethodBase": method_base,
                        "EndpointPath": rest,
                        "EndpointPath with version": full_path,
                        "SourceFormat": "MIN",
                        "VersionInference": "path-prefix or default v1",
                    },
                )
            )

        return Inventory(rows)

    def match(self, http_method: str, endpoint_path: str, version: Optional[str]) -> Tuple[List[InventoryRow], List[InventoryRow]]:
        hm = norm_http(http_method)
        ep = norm_endpoint_path(endpoint_path)
        v = norm_version(version) if version else None

        exact: List[InventoryRow] = []
        for r in self.rows:
            if r.http_method != hm:
                continue
            if r.endpoint_path != ep:
                continue
            if v and r.version != v:
                continue
            exact.append(r)

        suggestions: List[InventoryRow] = []
        ep_last = ep.split("/")[-1] if ep else ""
        for r in self.rows:
            if r.http_method != hm:
                continue
            if ep and (ep in r.endpoint_path or r.endpoint_path in ep):
                suggestions.append(r)
                continue
            if ep_last and ep_last == r.endpoint_path.split("/")[-1]:
                suggestions.append(r)

        return exact, suggestions

    def infer_version(self, http_method: str, endpoint_path: str) -> List[str]:
        hm = norm_http(http_method)
        ep = norm_endpoint_path(endpoint_path)
        return sorted({r.version for r in self.rows if r.http_method == hm and r.endpoint_path == ep})
