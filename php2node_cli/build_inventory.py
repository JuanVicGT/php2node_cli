from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


HTTP_SUFFIX = {
    "_get": "GET",
    "_post": "POST",
    "_put": "PUT",
    "_delete": "DELETE",
}


@dataclass
class EndpointRow:
    app: str  # api|portal
    version: str  # v1|v2|unknown
    controller: str  # bank_account
    http_method: str  # GET
    method: str  # dacustomer_bank_get
    method_base: str  # dacustomer_bank
    endpoint_path: str  # bank_account/dacustomer_bank
    endpoint_path_with_version: str  # v1/bank_account/dacustomer_bank
    controller_file: str
    class_name: Optional[str]
    notes: str


def norm(s: str) -> str:
    return (s or "").strip().replace("\\", "/")


def detect_app_and_rel(p: Path, repo_root: Path) -> Tuple[str, str]:
    """
    Devuelve app y ruta relativa normalizada.
    """
    rel = norm(str(p.relative_to(repo_root)))
    if rel.startswith("api/"):
        return "api", rel
    if rel.startswith("portal/"):
        return "portal", rel
    return "unknown", rel


def detect_version_from_path(rel: str) -> str:
    # Busca /controllers/v1/ o /controllers/v2/
    m = re.search(r"/controllers/(v\d+)/", rel, flags=re.IGNORECASE)
    if m:
        return m.group(1).lower()
    return "unknown"


def controller_name_from_file(p: Path) -> str:
    # CodeIgniter: bank_account.php -> bank_account
    return p.stem.lower()


def extract_class_name(text: str) -> Optional[str]:
    m = re.search(r"\bclass\s+([A-Za-z_][A-Za-z0-9_]*)\b", text)
    return m.group(1) if m else None


def find_methods_rest(text: str) -> List[str]:
    """
    Extrae nombres de métodos "function xxx_get(" etc.
    """
    # Nota: cubre "public function", "function", etc.
    candidates = re.findall(r"\bfunction\s+([A-Za-z_][A-Za-z0-9_]*)\s*\(", text)
    out: List[str] = []
    for name in candidates:
        lname = name.lower()
        for suffix in HTTP_SUFFIX.keys():
            if lname.endswith(suffix):
                out.append(name)
                break
    return out


def detect_notes(text: str) -> str:
    notes: List[str] = []
    if re.search(r"\bfunction\s+_remap\s*\(", text):
        notes.append("HAS__REMAP")
    if re.search(r"\broute\b", text, flags=re.IGNORECASE):
        # heurística débil, solo para flag
        notes.append("MAY_HAVE_CUSTOM_ROUTING")
    return ",".join(notes)


def iter_controller_files(repo_root: Path) -> Iterable[Path]:
    """
    Busca controllers en:
      api/application/controllers
      portal/application/controllers
    Incluye subcarpetas.
    """
    bases = [
        repo_root / "api" / "application" / "controllers",
        repo_root / "portal" / "application" / "controllers",
    ]
    for base in bases:
        if not base.exists():
            continue
        for p in base.rglob("*.php"):
            # evita rutas típicas que no son controllers (si aplica)
            yield p


def build_rows(repo_root: Path) -> List[EndpointRow]:
    rows: List[EndpointRow] = []

    for controller_file in iter_controller_files(repo_root):
        try:
            raw = controller_file.read_bytes()
        except Exception:
            continue

        # decode robusto
        text = None
        for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            text = raw.decode("latin-1", errors="replace")

        app, rel = detect_app_and_rel(controller_file, repo_root)
        version = detect_version_from_path(rel)
        controller = controller_name_from_file(controller_file)
        class_name = extract_class_name(text)
        notes = detect_notes(text)

        methods = find_methods_rest(text)
        for method in methods:
            m_lower = method.lower()
            http_method = None
            method_base = None
            for suffix, hm in HTTP_SUFFIX.items():
                if m_lower.endswith(suffix):
                    http_method = hm
                    method_base = method[: -len(suffix)]
                    break
            if not http_method or not method_base:
                continue

            ep = f"{controller}/{method_base}"
            if version != "unknown":
                epv = f"{version}/{ep}"
            else:
                epv = ep

            rows.append(
                EndpointRow(
                    app=app,
                    version=version,
                    controller=controller,
                    http_method=http_method,
                    method=method,
                    method_base=method_base,
                    endpoint_path=ep,
                    endpoint_path_with_version=epv,
                    controller_file=str(controller_file),
                    class_name=class_name,
                    notes=notes,
                )
            )

    # orden estable
    rows.sort(key=lambda r: (r.version, r.app, r.controller, r.http_method, r.method_base))
    return rows


def write_xlsx(rows: List[EndpointRow], out_path: Path, sheet_name: str = "EndPoints") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = [
        "Version",
        "App",
        "Controller",
        "HttpMethod",
        "Method",
        "MethodBase",
        "EndpointPath",
        "EndpointPath with version",
        "ControllerFile",
        "ClassName",
        "Notes",
    ]
    ws.append(headers)

    for r in rows:
        ws.append(
            [
                r.version,
                r.app,
                r.controller,
                r.http_method,
                r.method,
                r.method_base,
                r.endpoint_path,
                r.endpoint_path_with_version,
                r.controller_file,
                r.class_name or "",
                r.notes,
            ]
        )

    # autosize simple
    for col_idx, _ in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 80)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Build EndPoints inventory from CodeIgniter repo (REST_Controller style).")
    ap.add_argument("--repo-root", required=True, help="Root of repo (api-backend).")
    ap.add_argument("--out-xlsx", required=True, help="Output xlsx file path.")
    ap.add_argument("--sheet", default="EndPoints", help="Sheet name. Default: EndPoints")
    args = ap.parse_args()

    repo_root = Path(args.repo_root).resolve()
    out_xlsx = Path(args.out_xlsx).resolve()

    if not repo_root.exists():
        raise SystemExit(f"Repo root not found: {repo_root}")

    rows = build_rows(repo_root)
    write_xlsx(rows, out_xlsx, sheet_name=args.sheet)

    print(f"OK. Endpoints discovered: {len(rows)}")
    print(f"Output: {out_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())